#  El problema con este documento es que solo se pueden leer archivos que tengan el 
#  formato y estilo de titulos en las columnas, de igual manera sirvio para practicar.
import string    # se utiliza para obtener las columnas en excel (1=A, 2=B, etc.)
import pandas as pd # se utiliza para obtener los valores de excel. 
from openpyxl import load_workbook # Se generan el archivo xlsx
from openpyxl.chart import BarChart, Reference 
from openpyxl.styles import Font

def automatizar_excel(nombre_archivo):
    """ Input ventas_mes-xlsx / Output reporte_mes.xlsx"""

    archivo_excel = pd.read_excel(nombre_archivo)
    # print(archivo_excel[['Artículo', 'Color', 'Unidades']])
    tabla_pivote = archivo_excel.pivot_table(index='Artículo', columns='Color', values='Unidades', aggfunc='sum').round(0)
    # print(tabla_pivote)
    mes_extension = nombre_archivo.split('_')[1]
    tabla_pivote.to_excel(f'ventas_{mes_extension}.xlsx',startrow= 3, startcol=2, sheet_name='Reporte')

    wb = load_workbook(f'ventas_{mes_extension}.xlsx')
    pestaña = wb['Reporte']

    min_columna = wb.active.min_column
    max_columna = wb.active.max_column
    min_fila = wb.active.min_row
    max_fila = wb.active.max_row

    print(min_columna)
    print(max_columna)
    print(min_fila)
    print(max_fila)


    # totales de  las filas 
    abecedario = list(string.ascii_uppercase)
    abecedarioExcel = abecedario[min_columna-1:max_columna]
    print(abecedarioExcel)
    for i in abecedarioExcel:
        if i!='A':
            pestaña[f'{i}{max_fila+1}'] = f'=SUM({i}{min_fila+1}:{i}{max_fila})'
            # pestaña[f'{i}{max_fila+1}'].style = 'Currency'
            print(f'=SUM({i}{min_fila+1}:{i}{max_fila})')

    pestaña[f'{abecedarioExcel[0]}{max_fila+1}'] = 'Total_Nuevo'

    # Gráfico en Excel

    barchart = BarChart()

    data = Reference(pestaña, min_col=min_columna + 1, max_col=max_columna, min_row=min_fila, max_row=max_fila)

    categorias = Reference(pestaña, min_col=min_columna, max_col=min_columna, min_row=min_fila+1, max_row=max_fila+1)
    barchart.add_data(data, titles_from_data=True)
    barchart.set_categories(categorias)
    pestaña.add_chart(barchart, f'{abecedarioExcel[0]}{max_fila +3}')
    barchart.title ='Unidades'
    barchart.style = 5

    pestaña['A1'] = 'Reporte'
    mes = mes_extension.split('.')[0]
    pestaña['A2'] = mes

    pestaña['A1'].font = Font('Arial', bold=True, size=40)
    pestaña['A2'].font = Font('Arial', bold=True, size=12)

    wb.save(f'reporte_{mes_extension}')
    return





automatizar_excel('ventas_2021.xlsx')